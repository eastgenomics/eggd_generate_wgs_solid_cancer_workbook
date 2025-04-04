from openpyxl.styles import Border, Side
from openpyxl.styles.fills import PatternFill

from utils import html


# prepare formatting
THIN = Side(border_style="thin", color="000000")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
LOWER_BORDER = Border(bottom=THIN)

CONFIG = {
    "cells_to_write": {
        (4, 1): "Diagnosis Date",
        (4, 2): "Tumour Received",
        (4, 3): "Tumour ID",
        (4, 4): "Presentation",
        (4, 5): "Diagnosis",
        (4, 6): "Tumour Site",
        (4, 7): "Tumour Type",
        (4, 8): "Germline Sample",
        (5, 1): ("Tumor info", 0, "Tumour Diagnosis Date"),
        (5, 2): ("Sample info", 0, "Clinical Sample Date Time"),
        (5, 3): ("Tumor info", 0, "Histopathology or SIHMDS LAB ID"),
        (5, 4): [
            ("Tumor info", 0, "Presentation", "split"),
            ("Tumor info", 0, "Primary or Metastatic", "parentheses"),
        ],
        (5, 5): ("Patient info", 0, "Clinical Indication"),
        (5, 6): ("Tumor info", 0, "Tumour Topography"),
        (5, 7): [
            ("Sample info", 0, "Storage Medium", ""),
            ("Sample info", 0, "Source", ""),
        ],
        (5, 8): [
            ("Germline info", 0, "Storage Medium", ""),
            ("Germline info", 0, "Source", "parentheses"),
        ],
        (7, 1): "Purity (Histo)",
        (7, 2): "Purity (Calculated)",
        (7, 3): "Ploidy",
        (7, 4): "Total SNVs",
        (7, 5): "Total Indels",
        (7, 6): "Total SVs",
        (7, 7): "TMB",
        (8, 1): ("Sample info", 0, "Tumour Content"),
        (8, 2): ("Sample info", 0, "Calculated Tumour Content"),
        (8, 3): ("Sample info", 0, "Calculated Overall Ploidy"),
        (8, 4): ("Sequencing info", 1, "Total somatic SNVs"),
        (8, 5): ("Sequencing info", 1, "Total somatic indels"),
        (8, 6): ("Sequencing info", 1, "Total somatic SVs"),
        (8, 7): html.get_tag_sibling,
        (10, 1): "Sample type",
        (10, 2): "Mean depth, x",
        (10, 3): "Mapped reads, %",
        (10, 4): "Chimeric DNA frag, %",
        (10, 5): "Insert size, bp",
        (10, 6): "Unevenness, x",
        (11, 1): ("Sequencing info", 0, "Sample type"),
        (11, 2): (
            "Sequencing info",
            0,
            "Genome-wide coverage mean, x",
        ),
        (11, 3): ("Sequencing info", 0, "Mapped reads, %"),
        (11, 4): ("Sequencing info", 0, "Chimeric DNA fragments, %"),
        (11, 5): ("Sequencing info", 0, "Insert size median, bp"),
        (11, 6): (
            "Sequencing info",
            0,
            "Unevenness of local genome coverage, x",
        ),
        (12, 1): ("Sequencing info", 1, "Sample type"),
        (12, 2): (
            "Sequencing info",
            1,
            "Genome-wide coverage mean, x",
        ),
        (12, 3): ("Sequencing info", 1, "Mapped reads, %"),
        (12, 4): ("Sequencing info", 1, "Chimeric DNA fragments, %"),
        (12, 5): ("Sequencing info", 1, "Insert size median, bp"),
        (12, 6): (
            "Sequencing info",
            1,
            "Unevenness of local genome coverage, x",
        ),
        (1, 1): "=SOC!A2",
        (2, 1): "=SOC!A3",
        (1, 3): "=SOC!A5",
        (2, 3): "=SOC!A6",
        (1, 5): "=SOC!A9",
        (15, 1): "QC alerts",
        (16, 1): "None",
        (15, 2): "Assessed purity",
        (15, 3): "SNV TMB",
    },
    "to_bold": [
        "A1",
        "A4",
        "A7",
        "A10",
        "A15",
        "B4",
        "B7",
        "B10",
        "C4",
        "C7",
        "C10",
        "D4",
        "D7",
        "D10",
        "E4",
        "E7",
        "E10",
        "F4",
        "F7",
        "F10",
        "G4",
        "G7",
        "H4",
    ],
    "col_width": [
        ("A", 22),
        ("B", 22),
        ("C", 22),
        ("D", 22),
        ("E", 22),
        ("F", 22),
        ("G", 22),
        ("H", 22),
        ("I", 22),
        ("J", 22),
    ],
    "cells_to_colour": [
        ("A4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("B4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("C4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("D4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("E4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("F4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("G4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("H4", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("A7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("B7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("C7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("D7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("E7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("F7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("G7", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("A10", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("B10", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("C10", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("D10", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("E10", PatternFill(patternType="solid", start_color="ADD8E6")),
        ("F10", PatternFill(patternType="solid", start_color="ADD8E6")),
    ],
    "borders": {
        "single_cells": [
            ("A15", LOWER_BORDER),
        ],
        "cell_rows": [
            ("A4:H4", THIN_BORDER),
            ("A5:H5", THIN_BORDER),
            ("A7:G7", THIN_BORDER),
            ("A8:G8", THIN_BORDER),
            ("A10:F10", THIN_BORDER),
            ("A11:F11", THIN_BORDER),
            ("A12:F12", THIN_BORDER),
        ],
    },
    "dropdowns": [
        {
            "cells": {
                ("A16",): (
                    '"None,<30% tumour purity,SNVs low VAF (<6%),TINC (<5%),'
                    'Somatic CNV, Germline CNV"'
                ),
            },
            "title": "QC alerts",
        }
    ],
    "images": [
        {"cell": "A18", "img_index": 8, "size": (600, 800)},
        {"cell": "H18", "img_index": 10, "size": (600, 800)},
    ],
}
