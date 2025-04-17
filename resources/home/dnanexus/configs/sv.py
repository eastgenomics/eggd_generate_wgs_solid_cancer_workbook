import string

from openpyxl.styles.fills import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from utils import misc


CONFIG = {
    "col_width": [
        ("B", 12),
        ("C", 22),
        ("D", 22),
        ("E", 20),
        ("G", 16),
        ("H", 16),
        ("I", 18),
        ("J", 18),
        ("K", 18),
        ("L", 18),
        ("M", 18),
        ("N", 18),
        ("O", 18),
    ],
    "freeze_panes": "F1",
    "expected_columns": [
        "Event domain",
        "Impacted transcript region",
        "Gene",
        "GRCh38 coordinates",
        "Chromosomal bands",
        "Type",
        "Fusion",
        "Size",
        (
            "Population germline allele frequency (GESG | GECG for somatic "
            "SVs or AF | AUC for germline CNVs)",
        ),
        "Paired reads",
        "Split reads",
        "Gene mode of action",
        "Variant class",
        "Actionability",
        "Comments",
    ],
    "alternative_columns": [
        [
            (
                "Population germline allele frequency (GESG | GECG for "
                "somatic SVs or AF | AUC for germline CNVs)",
            ),
            (
                "Population germline allele frequency (AF | AUC for germline "
                "CNVs)"
            ),
        ]
    ],
}


def add_dynamic_values(data: pd.DataFrame) -> dict:
    """Add dynamic values for the SV sheet

    Parameters
    ----------
    data : pd.DataFrame
        Dataframe containing the data for SV fusion variants and appropriate
        additional data from inputs

    Returns
    -------
    dict
        Dict containing data that needs to be merged to the CONFIG variable
    """

    nb_structural_variants = data.shape[0]

    last_column_letter = misc.get_column_letter_using_column_name(data)
    variant_class_column_letter = misc.get_column_letter_using_column_name(
        data, "Variant class"
    )

    lookup_groups = misc.get_lookup_groups(data)

    cells_to_color = []

    # build the cells to color data
    for i, index_group in enumerate(lookup_groups):
        for index in index_group:
            for j in range(1, nb_structural_variants + 2):
                if i % 2 == 0:
                    pattern = PatternFill(
                        patternType="solid", start_color="c4d9ef"
                    )
                else:
                    pattern = PatternFill(
                        patternType="solid", start_color="B8E7E0"
                    )

                cells_to_color.append(
                    (f"{string.ascii_uppercase[index]}{j}", pattern)
                )

    config_with_dynamic_values = {
        "cells_to_write": {
            (1, i): column for i, column in enumerate(data.columns, 1)
        }
        | {
            # remove the col and row index from the writing?
            (r_idx - 1, c_idx - 1): value
            for r_idx, row in enumerate(dataframe_to_rows(data), 1)
            for c_idx, value in enumerate(row, 1)
            if c_idx != 1 and r_idx != 1
        },
        "cells_to_colour": [
            (
                f"{variant_class_column_letter}{i}",
                PatternFill(patternType="solid", start_color="FFDBBB"),
            )
            for i in range(1, nb_structural_variants + 2)
        ]
        + cells_to_color,
        "to_bold": [
            f"{string.ascii_uppercase[i]}1"
            for i in range(
                misc.convert_letter_column_to_index(last_column_letter)
            )
        ],
        "dropdowns": [
            {
                "cells": {
                    (
                        f"{variant_class_column_letter}{i}"
                        for i in range(2, nb_structural_variants + 2)
                    ): (
                        '"Oncogenic, Likely oncogenic,'
                        "Uncertain, Likely passenger,"
                        'Likely artefact"'
                    ),
                },
                "title": "Variant class",
            },
        ],
        "auto_filter": f"F:{last_column_letter}",
    }

    return config_with_dynamic_values
