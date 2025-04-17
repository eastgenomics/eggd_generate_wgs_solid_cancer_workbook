from bs4 import BeautifulSoup
import openpyxl
from openpyxl import drawing
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from PIL import Image

from configs.tables import get_table_value_in_html_table
from utils import misc

pd.options.mode.chained_assignment = None


def write_sheet(
    excel_writer: pd.ExcelWriter,
    sheet_name: str,
    html_tables: list = None,
    html_images: list = None,
    soup: BeautifulSoup = None,
    dynamic_data: dict = None,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Using a config file, write in the appropriate data

    Parameters
    ----------
    excel_writer : pd.ExcelWriter
        ExcelWriter object
    sheet_name : str
        Name of the sheet used to match the config
    html_tables : list, optional
        List of tables extracted from the HTML
    html_images : list, optional
        List of images extracted from the HTML
    soup : BeautifulSoup, optional
        BeautifulSoup object for the HTML file
    dynamic_data: dict, optional
        Dict of data for dynamic filling in the sheet

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        Worksheet object
    """

    sheet = excel_writer.book.create_sheet(sheet_name)

    type_config = misc.select_config(sheet_name)
    assert type_config, f"Config file {sheet_name} couldn't be imported"

    if dynamic_data:
        dynamic_data_for_config = dynamic_data[sheet_name]
        sheet_config = misc.merge_dicts(
            type_config.CONFIG, dynamic_data_for_config
        )

    else:
        sheet_config = type_config.CONFIG

    if sheet_config.get("cells_to_write"):
        write_cell_content(
            sheet, sheet_config["cells_to_write"], html_tables, soup
        )

    if sheet_config.get("to_merge"):
        # merge columns that have longer text
        sheet.merge_cells(**sheet_config["to_merge"])

    if sheet_config.get("to_align"):
        align_cells(sheet, sheet_config["to_align"])

    if sheet_config.get("to_bold"):
        bold_cells(sheet, sheet_config["to_bold"])

    if sheet_config.get("col_width"):
        set_col_width(sheet, sheet_config["col_width"])

    if sheet_config.get("row_height"):
        set_row_height(sheet, sheet_config["row_height"])

    if sheet_config.get("cells_to_colour"):
        color_cells(sheet, sheet_config["cells_to_colour"])

    if sheet_config.get("borders"):
        draw_borders(sheet, sheet_config["borders"])

    if sheet_config.get("dropdowns"):
        generate_dropdowns(sheet, sheet_config["dropdowns"])

    if sheet_config.get("images"):
        insert_images(sheet, sheet_config["images"], html_images)

    if sheet_config.get("auto_filter"):
        filters = sheet.auto_filter
        filters.ref = sheet_config["auto_filter"]

    if sheet_config.get("freeze_panes"):
        sheet.freeze_panes = sheet[sheet_config["freeze_panes"]]

    if sheet_config.get("data_bar"):
        add_databar_rule(sheet, sheet_config["data_bar"])

    return sheet


def write_cell_content(
    sheet: Worksheet, config_data: dict, html_tables: list, soup: BeautifulSoup
):
    """Write the tables from the config

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to write the tables into
    config_data : dict
        Dict of tables to write
    html_tables: list
        List of dict for the tables extracted from the HTML
    soup: BeautifulSoup
        HTML page
    """

    for cell_pos, value in config_data.items():
        cell_x, cell_y = cell_pos

        if type(value) in [str, float, int]:
            value_to_write = value

        elif type(value) is list:
            value_to_write = []

            for (
                table_name_in_config,
                row,
                column,
                formatting,
            ) in value:
                subvalue = get_table_value_in_html_table(
                    table_name_in_config,
                    row,
                    column,
                    html_tables,
                    formatting,
                )
                value_to_write.append(subvalue)

            value_to_write = " ".join(value_to_write)

        # single value to add in the table
        elif type(value) is tuple:
            table_name_in_config, row, column = value
            value_to_write = get_table_value_in_html_table(
                table_name_in_config, row, column, html_tables
            )

        elif value is None:
            value_to_write = ""

        else:
            # special hardcoded case, haven't found a way to make that
            # better for now (which means it'll probably stay that way
            # forever)
            value_to_write = value(
                soup,
                "b",
                (
                    "Total number of somatic non-synonymous small "
                    "variants per megabase"
                ),
            )

        sheet.cell(cell_x, cell_y).value = value_to_write


def align_cells(sheet: Worksheet, config_data: list):
    """For given list of cells, align the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to align the cells
    config_data : list
        List of cells to align
    """

    for cell in config_data:
        sheet[cell].alignment = Alignment(wrapText=True, horizontal="center")


def bold_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells, bold them

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to bold the cells
    config_data : list
        List of cells to bold
    """

    for cell in config_data:
        sheet[cell].font = Font(bold=True, name=DEFAULT_FONT.name)


def set_col_width(sheet: Worksheet, config_data: list):
    """Given a list of columns, set their width

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to set the width
    config_data : list
        List of tuple with the column and its width to set
    """

    for cell, width in config_data:
        sheet.column_dimensions[cell].width = width


def set_row_height(sheet: Worksheet, config_data: list):
    """Given a list of rows, set their height

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to set the height
    config_data : list
        List of tuple with the row and its height to set
    """

    for cell, height in config_data:
        sheet.row_dimensions[cell].height = height


def color_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells and their color, color the cells appropriately

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to color the cells in
    config_data : list
        List of tuples with the cells and their color
    """

    for cell, color in config_data:
        sheet[cell].fill = color


def draw_borders(sheet: Worksheet, config_data: dict):
    """Draw borders around the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to draw borders
    config_data : dict
        Dict containing info for the single cells to draw borders around and
        the rows of cells
    """

    if config_data.get("single_cells"):
        for cell, type_border in config_data["single_cells"]:
            sheet[cell].border = type_border

    if config_data.get("cell_rows"):
        for cell_range, type_border in config_data["cell_rows"]:
            for cells in sheet[cell_range]:
                for cell in cells:
                    cell.border = type_border


def generate_dropdowns(sheet: Worksheet, config_data: dict):
    """Write in the dropdown menus

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the dropdown menus
    config_data : dict
        Dict of data for the dropdown menus
    """

    for dropdown_info in config_data:
        for cells, options in dropdown_info["cells"].items():
            dropdown = DataValidation(
                type="list", formula1=options, allow_blank=True
            )
            dropdown.prompt = "Select from the list"
            dropdown.promptTitle = dropdown_info["title"]
            dropdown.showInputMessage = True
            dropdown.showErrorMessage = True
            sheet.add_data_validation(dropdown)

            for cell in cells:
                dropdown.add(sheet[cell])


def insert_images(sheet: Worksheet, config_data: dict, images: list):
    """Insert images in the given worksheet for that config file

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the images
    config_data : list
        List of image data
    images: list
        List of images extracted from the HTML file
    """

    for image_data in config_data:
        height, width = image_data["size"]
        image_pil_obj = Image.open(images[image_data["img_index"]])
        image = drawing.image.Image(image_pil_obj)
        image.height = height
        image.width = width
        image.anchor = image_data["cell"]
        sheet.add_image(image)


def add_databar_rule(sheet: Worksheet, range_cell: str):
    """Add a databar for the range of cells given

    Parameters
    ----------
    sheet : Worksheet
        Sheet to add the databar(s) to
    range_cell : str
        String in "COL#:COL#" format for position of databar(s)
    """

    sheet.conditional_formatting.add(
        range_cell,
        DataBarRule(
            start_type="num",
            start_value=0,
            end_type="num",
            end_value=1,
            color="FF3361",
        ),
    )
