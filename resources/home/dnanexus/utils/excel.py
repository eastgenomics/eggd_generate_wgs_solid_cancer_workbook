from bs4 import BeautifulSoup
import openpyxl
from openpyxl import drawing
from openpyxl.styles import Alignment, DEFAULT_FONT, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
import pandas as pd
from PIL import Image
import vcfpy

from configs.tables import get_table_value_in_html_table
from utils import misc, vcf


def open_file(file: str, file_type: str) -> pd.DataFrame:
    """Read in CSV or XLS files using pandas

    Parameters
    ----------
    file : str
        File path
    file_type : str
        File type with the file path

    Returns
    -------
    pd.DataFrame
        Dataframe created by pandas
    """

    if file_type == "csv":
        df = pd.read_csv(file)
    elif file_type == "xls":
        df = pd.read_excel(file)

    if "ClinVar ID" in df.columns:
        df["ClinVar ID"] = df["ClinVar ID"].astype(str)
        df["ClinVar ID"] = df["ClinVar ID"].str.removesuffix(".0")

    return df


def process_reported_variants_germline(
    df: pd.DataFrame, clinvar_resource: vcfpy.Reader
) -> pd.DataFrame:
    """Process the data from the reported variants excel file

    Parameters
    ----------
    df : pd.DataFrame
        Dataframe from parsing the reported variants excel file
    clinvar_resource : vcfpy.Reader
        vcfpy.Reader object from the Clinvar resource

    Returns
    -------
    pd.DataFrame
        Dataframe containing clinical significance info for germline variants
    """

    df = df[df["Origin"].str.lower() == "germline"]

    if df.empty:
        return None

    df.reset_index(drop=True, inplace=True)

    clinvar_ids_to_find = [
        value for value in df.loc[:, "ClinVar ID"].to_numpy()
    ]
    clinvar_info = vcf.find_clinvar_info(
        clinvar_resource, *clinvar_ids_to_find
    )

    # add the clinvar info by merging the clinvar dataframe
    df = df.merge(clinvar_info, on="ClinVar ID", how="left")

    # split the col to get gnomAD
    df[["GE", "gnomAD"]] = df[
        "Population germline allele frequency (GE | gnomAD)"
    ].str.split("|", expand=True)

    df.drop(
        ["GE", "Population germline allele frequency (GE | gnomAD)"],
        axis=1,
        inplace=True,
    )
    df.loc[:, "Variant Class"] = ""
    df.loc[:, "Actionability"] = ""
    df = df[
        [
            "Gene",
            "GRCh38 coordinates;ref/alt allele",
            "CDS change and protein change",
            "Predicted consequences",
            "Genotype",
            "Variant Class",
            "Actionability",
            "Gene mode of action",
            "clnsigconf",
            "gnomAD",
        ]
    ]

    return df


def process_reported_variants_somatic(df: pd.DataFrame) -> pd.DataFrame:
    """Get the somatic variants and format the data for them

    Parameters
    ----------
    df : pd.DataFrame
        Dataframe from parsing the reported variants excel file

    Returns
    -------
    pd.DataFrame
        Dataframe with additional formatting for c. and p. annotation
    """

    # select only somatic rows
    df = df[df["Origin"].str.lower().str.contains("somatic")]
    df.reset_index(drop=True, inplace=True)
    df[["c_dot", "p_dot"]] = df["CDS change and protein change"].str.split(
        r"(?=;p)", n=1, expand=True
    )
    df["p_dot"] = df["p_dot"].str.slice(1)

    return df


def write_sheet(
    excel_writer: pd.ExcelWriter,
    sheet_name: str,
    html_tables: list = None,
    html_images: list = None,
    soup: BeautifulSoup = None,
    sheet_data: dict = None,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Using a config file, write in the appropriate data

    Parameters
    ----------
    excel_writer : pd.ExcelWriter
        ExcelWriter object
    sheet_name : str
        Name of the sheet used to match the config
    html_tables : list, optional
        List of tables extracted from the HTML
    html_images : list, optional
        List of images extracted from the HTML
    soup : BeautifulSoup, optional
        BeautifulSoup object for the HTML file
    sheet_data: dict, optional
        Dict of data for dynamic filling in the sheet

    Returns
    -------
    openpyxl.worksheet.worksheet.Worksheet
        Worksheet object
    """

    sheet = excel_writer.book.create_sheet(sheet_name)

    type_config = misc.select_config(sheet_name)
    assert type_config, "Config file couldn't be imported"

    if hasattr(type_config, "fill_values"):
        sheet_config = type_config.fill_values(sheet_data)
    else:
        sheet_config = type_config.CONFIG

    if sheet_config.get("tables"):
        write_tables(sheet, sheet_config["tables"], html_tables, soup)

    if sheet_config.get("to_merge"):
        # merge columns that have longer text
        sheet.merge_cells(**sheet_config["to_merge"])

    if sheet_config.get("to_align"):
        align_cells(sheet, sheet_config["to_align"])

    if sheet_config.get("to_bold"):
        bold_cells(sheet, sheet_config["to_bold"])

    if sheet_config.get("col_width"):
        set_col_width(sheet, sheet_config["col_width"])

    if sheet_config.get("cells_to_colour"):
        color_cells(sheet, sheet_config["cells_to_colour"])

    if sheet_config.get("borders"):
        draw_borders(sheet, sheet_config["borders"])

    if sheet_config.get("dropdowns"):
        generate_dropdowns(sheet, sheet_config["dropdowns"])

    if sheet_config.get("images"):
        insert_images(sheet, sheet_config["images"], html_images)

    return sheet


def write_tables(
    sheet: Worksheet, config_data: list, html_tables: list, soup: BeautifulSoup
):
    """Write the tables from the config

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to write the tables into
    config_data : list
        List of tables to write
    html_tables: list
        List of dict for the tables extracted from the HTML
    soup: BeautifulSoup
        HTML page
    """

    for config_table in config_data:
        headers = config_table["headers"]

        for cell_x, cell_y in headers:
            value_to_write = headers[cell_x, cell_y]
            sheet.cell(cell_x, cell_y).value = value_to_write

        if config_table.get("values"):
            values = config_table.get("values")

            for cell_x, cell_y in values:
                # if the value is a list, it means that concatenation is
                # required
                if isinstance(values[cell_x, cell_y], list):
                    value_to_write = []

                    for (
                        table_name_in_config,
                        row,
                        column,
                        formatting,
                    ) in values[cell_x, cell_y]:
                        subvalue = get_table_value_in_html_table(
                            table_name_in_config,
                            row,
                            column,
                            html_tables,
                            formatting,
                        )
                        value_to_write.append(subvalue)

                    value_to_write = " ".join(value_to_write)

                # single value to add in the table
                elif isinstance(values[cell_x, cell_y], tuple):
                    table_name_in_config, row, column = values[cell_x, cell_y]
                    value_to_write = get_table_value_in_html_table(
                        table_name_in_config, row, column, html_tables
                    )
                else:
                    # special hardcoded case, haven't found a way to make that
                    # better for now (which means it'll probably stay that way
                    # forever)
                    value_to_write = values[cell_x, cell_y](
                        soup,
                        "b",
                        (
                            "Total number of somatic non-synonymous small "
                            "variants per megabase"
                        ),
                    )

                sheet.cell(cell_x, cell_y).value = value_to_write

        if config_table.get("dynamic_values"):
            dynamic_values = config_table.get("dynamic_values")

            for cell_x, cell_y in dynamic_values:
                value_to_write = dynamic_values[cell_x, cell_y]
                sheet.cell(cell_x, cell_y).value = value_to_write


def align_cells(sheet: Worksheet, config_data: list):
    """For given list of cells, align the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to align the cells
    config_data : list
        List of cells to align
    """

    for cell in config_data:
        sheet[cell].alignment = Alignment(wrapText=True, horizontal="center")


def bold_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells, bold them

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to bold the cells
    config_data : list
        List of cells to bold
    """

    for cell in config_data:
        sheet[cell].font = Font(bold=True, name=DEFAULT_FONT.name)


def set_col_width(sheet: Worksheet, config_data: list):
    """Given a list of columns, set their width

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to set the width
    config_data : list
        List of tuple with the column and its width to set
    """

    for cell, width in config_data:
        sheet.column_dimensions[cell].width = width


def color_cells(sheet: Worksheet, config_data: list):
    """Given a list of cells and their color, color the cells appropriately

    Parameters
    ----------
    sheet : Worksheet
        Worksheet to color the cells in
    config_data : list
        List of tuples with the cells and their color
    """

    for cell, color in config_data:
        sheet[cell].fill = color


def draw_borders(sheet: Worksheet, config_data: dict):
    """Draw borders around the cells

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to draw borders
    config_data : dict
        Dict containing info for the single cells to draw borders around and
        the rows of cells
    """

    if config_data.get("single_cells"):
        for cell, type_border in config_data["single_cells"]:
            sheet[cell].border = type_border

    if config_data.get("cell_rows"):
        for cell_range, type_border in config_data["cell_rows"]:
            for cells in sheet[cell_range]:
                for cell in cells:
                    cell.border = type_border


def generate_dropdowns(sheet: Worksheet, config_data: dict):
    """Write in the dropdown menus

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the dropdown menus
    config_data : dict
        Dict of data for the dropdown menus
    """

    for cells, options in config_data["cells"].items():
        dropdown = DataValidation(
            type="list", formula1=options, allow_blank=True
        )
        dropdown.prompt = "Select from the list"
        dropdown.promptTitle = config_data["title"]
        dropdown.showInputMessage = True
        dropdown.showErrorMessage = True
        sheet.add_data_validation(dropdown)

        for cell in cells:
            dropdown.add(sheet[cell])


def insert_images(sheet: Worksheet, config_data: dict, images: list):
    """Insert images in the given worksheet for that config file

    Parameters
    ----------
    sheet : Worksheet
        Worksheet in which to write the images
    config_data : list
        List of image data
    images: list
        List of images extracted from the HTML file
    """

    for image_data in config_data:
        height, width = image_data["size"]
        image_pil_obj = Image.open(images[image_data["img_index"]])
        image = drawing.image.Image(image_pil_obj)
        image.height = height
        image.width = width
        image.anchor = image_data["cell"]
        sheet.add_image(image)
